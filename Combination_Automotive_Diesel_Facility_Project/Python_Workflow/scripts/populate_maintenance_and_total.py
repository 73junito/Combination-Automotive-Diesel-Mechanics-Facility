#!/usr/bin/env python3
"""Populate Maintenance and Total Facility Cost PDFs.

Reads existing CSVs/PDFs for subtotals and writes two PDFs into scripts/outputs.
"""

import csv
import os
from decimal import Decimal, InvalidOperation

ROOT = os.path.dirname(__file__)
OUT = os.path.join(ROOT, "outputs")
SRC = os.path.abspath(os.path.join(ROOT, "..", "outputs"))
os.makedirs(OUT, exist_ok=True)


def sum_csv_totalcost(path):
    total = Decimal("0")
    if not os.path.exists(path):
        return None
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        header = next(r, None)
        # find TotalCost column
        tc_idx = None
        if header:
            for i, h in enumerate(header):
                if h and "total" in h.lower():
                    tc_idx = i
                    break
        for row in r:
            if tc_idx is not None and tc_idx < len(row):
                try:
                    total += Decimal(str(row[tc_idx]))
                except (InvalidOperation, ValueError):
                    pass
    return total


def extract_subtotal_from_pdf(path):
    try:
        from pypdf import PdfReader
    except ImportError:
        return None
    if not os.path.exists(path):
        return None
    try:
        reader = PdfReader(path)
        text = ""
        for p in reader.pages:
            text += (p.extract_text() or "") + "\n"
        import re

        m = re.search(r"Subtotal:\s*\$?([0-9,]+\.?[0-9]*)", text, re.IGNORECASE)
        if m:
            s = m.group(1).replace(",", "")
            return Decimal(s)
    except Exception as e:
        return None
    return None


def make_pdf_table(title, rows, subtotal, outpath):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    doc = SimpleDocTemplate(
        outpath,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    elems = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    t = Table(rows, colWidths=[320, 140])
    t.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elems.append(t)
    elems.append(Spacer(1, 12))
    elems.append(Paragraph(f"Subtotal: ${subtotal:,.2f}", styles["Normal"]))
    elems.append(Spacer(1, 6))
    elems.append(
        Paragraph(
            "Maintenance and replacement costs are estimated annual planning-level values.",
            styles["Italic"],
        )
    )
    doc.build(elems)


def make_total_pdf(outpath, lines, grand_total):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    doc = SimpleDocTemplate(
        outpath,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    elems = [Paragraph("Total_Facility_Cost_Estimate", styles["Title"]), Spacer(1, 12)]
    t = Table(lines, colWidths=[360, 140])
    t.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elems.append(t)
    elems.append(Spacer(1, 12))
    elems.append(Paragraph(f"Grand total: ${grand_total:,.2f}", styles["Normal"]))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph("Costs are planning-level estimates.", styles["Italic"]))
    doc.build(elems)


def main():
    ess_path = os.path.join(SRC, "essential_equipment.csv")
    non_path = os.path.join(SRC, "nonessential_equipment.csv")
    equip_xlsx = os.path.join(SRC, "equipment_lists.xlsx")

    ess_total = sum_csv_totalcost(ess_path) or Decimal("0")
    non_total = sum_csv_totalcost(non_path) or Decimal("0")

    # attempt furniture subtotal from equipment_lists.xlsx by reading 'TotalCost' column
    furn_total = None
    if os.path.exists(equip_xlsx):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(equip_xlsx, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            header = [str(c).strip() for c in next(ws.iter_rows(values_only=True))]
            tc_idx = None
            for i, h in enumerate(header):
                if "total" in h.lower():
                    tc_idx = i
                    break
            total = Decimal("0")
            if tc_idx is not None:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        v = r[tc_idx]
                        if v is None:
                            continue
                        total += Decimal(str(v))
                    except (InvalidOperation, ValueError):
                        pass
                furn_total = total
        except ImportError:
            furn_total = None

    # fallback: try extract from furniture PDF
    if furn_total is None:
        furn_pdf = os.path.join(OUT, "Furniture_List_with_Cost.pdf")
        furn_total = extract_subtotal_from_pdf(furn_pdf) or Decimal("0")

    # build maintenance items (conservative planning-level estimates)
    maintenance_items = [
        ("Filters & consumables", Decimal("1200")),
        ("Belts & hoses", Decimal("1000")),
        ("Fluids (oil/coolant)", Decimal("2000")),
        ("Inspections & calibrations", Decimal("1500")),
        ("Minor replacements", Decimal("2000")),
        ("Misc consumables", Decimal("800")),
    ]
    maint_sub = sum(v for _, v in maintenance_items)

    # write maintenance pdf
    rows = [["Item", "Annual Cost"]] + [[n, f"${v:,.2f}"] for n, v in maintenance_items]
    maint_out = os.path.join(OUT, "Maintenance_and_Replacement_Costs.pdf")
    make_pdf_table("Maintenance_and_Replacement_Costs", rows, maint_sub, maint_out)

    # write total pdf
    lines = [["Category", "Cost"]] + [
        ["Essential equipment subtotal", f"${ess_total:,.2f}"],
        ["Non-essential equipment subtotal", f"${non_total:,.2f}"],
        ["Furniture subtotal", f"${furn_total:,.2f}"],
        ["Maintenance (annual)", f"${maint_sub:,.2f}"],
    ]
    grand = ess_total + non_total + furn_total + maint_sub
    total_out = os.path.join(OUT, "Total_Facility_Cost_Estimate.pdf")
    make_total_pdf(total_out, lines, grand)

    print("Maintenance subtotal:", f"${maint_sub:,.2f}")
    print("Grand total:", f"${grand:,.2f}")


if __name__ == "__main__":
    main()
