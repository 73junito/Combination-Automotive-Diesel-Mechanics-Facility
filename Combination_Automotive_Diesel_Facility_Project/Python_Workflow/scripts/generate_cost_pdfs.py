#!/usr/bin/env python3
"""Generate instructor-ready PDFs from CSV/XLSX cost tables.

Creates these PDFs in scripts/outputs:
 - Essential_Equipment_Cost_List.pdf
 - Nonessential_Equipment_Cost_List.pdf
 - Furniture_List_with_Cost.pdf
 - Maintenance_and_Replacement_Costs.pdf
 - Total_Facility_Cost_Estimate.pdf

Behavior:
 - Prefer CSV files in ../outputs (same folder used by other scripts).
 - If CSV missing, try XLSX (openpyxl).
 - If no source, generate a placeholder PDF with guidance.
"""

import csv
import os
import sys
from decimal import Decimal, InvalidOperation

OUT = os.path.join(os.path.dirname(__file__), "outputs")
SRC = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "outputs"))
os.makedirs(OUT, exist_ok=True)

FILES = [
    (
        "essential_equipment.csv",
        "essential_equipment.xlsx",
        "Essential_Equipment_Cost_List.pdf",
    ),
    (
        "nonessential_equipment.csv",
        "nonessential_equipment.xlsx",
        "Nonessential_Equipment_Cost_List.pdf",
    ),
    ("furniture.csv", "equipment_lists.xlsx", "Furniture_List_with_Cost.pdf"),
    (
        "maintenance.csv",
        "maintenance_and_replacement.csv",
        "Maintenance_and_Replacement_Costs.pdf",
    ),
    ("totals.csv", "total_costs.xlsx", "Total_Facility_Cost_Estimate.pdf"),
]


def read_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        for row in r:
            rows.append(row)
    return rows


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows = []
    for r in sheet.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in r])
    return rows


def find_source(csv_name, xlsx_name):
    csvp = os.path.join(SRC, csv_name)
    if os.path.exists(csvp):
        return ("csv", csvp)
    xlsxp = os.path.join(SRC, xlsx_name)
    if os.path.exists(xlsxp):
        return ("xlsx", xlsxp)
    return (None, None)


def try_parse_money(s):
    try:
        s2 = str(s).replace("$", "").replace(",", "").strip()
        return Decimal(s2)
    except (InvalidOperation, ValueError):
        return None


def build_table_rows(rows):
    # Expect header row with columns like Category,Item,Quantity,UnitCost,TotalCost
    if not rows:
        return ([], Decimal(0))
    header = rows[0]
    # find useful columns
    col_map = {h.lower(): i for i, h in enumerate(header)}
    item_i = col_map.get("item", 0)
    qty_i = col_map.get("quantity", None)
    unit_i = col_map.get("unitcost", None)
    total_i = col_map.get("totalcost", None)

    table = [["Item", "Qty", "Unit Cost", "Total"]]
    subtotal = Decimal(0)
    for r in rows[1:]:
        item = r[item_i] if item_i < len(r) else ""
        qty = r[qty_i] if (qty_i is not None and qty_i < len(r)) else ""
        unit = r[unit_i] if (unit_i is not None and unit_i < len(r)) else ""
        total = r[total_i] if (total_i is not None and total_i < len(r)) else ""
        # try compute total if missing
        if (not total) and unit and qty:
            try:
                q = Decimal(str(qty))
                u = try_parse_money(unit) or Decimal(0)
                tot = q * u
                total = f"{tot:.2f}"
            except (InvalidOperation, TypeError, ValueError):
                total = ""
        val = try_parse_money(total)
        if val is not None:
            subtotal += val
            total_str = f"${val:,.2f}"
        else:
            total_str = str(total)

        unit_val = try_parse_money(unit)
        unit_str = f"${unit_val:,.2f}" if unit_val is not None else str(unit)
        table.append([str(item), str(qty), unit_str, total_str])
    return (table, subtotal)


def make_pdf(title, table_rows, subtotal, out_path):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)
    except ImportError as e:
        print("Missing ReportLab. Install reportlab in your Python environment.")
        raise

    doc = SimpleDocTemplate(
        out_path,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(title, styles["Title"]))
    elems.append(Spacer(1, 12))

    t = Table(table_rows, colWidths=[240, 60, 90, 90])
    t.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elems.append(t)
    elems.append(Spacer(1, 12))
    elems.append(Paragraph(f"Subtotal: ${subtotal:,.2f}", styles["Normal"]))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph("Planning-level estimate.", styles["Italic"]))
    doc.build(elems)


def make_placeholder(title, out_path, note=None):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
    except ImportError:
        print("Missing ReportLab. Install reportlab in your Python environment.")
        raise
    doc = SimpleDocTemplate(out_path, pagesize=letter)
    styles = getSampleStyleSheet()
    elems = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    elems.append(
        Paragraph(
            "Source data not found. Please provide the CSV or XLSX with the expected name.",
            styles["Normal"],
        )
    )
    if note:
        elems.append(Spacer(1, 6))
        elems.append(Paragraph(note, styles["Italic"]))
    doc.build(elems)


def main():
    for csvn, xlsn, outname in FILES:
        kind, path = find_source(csvn, xlsn)
        outpath = os.path.join(OUT, outname)
        if not kind:
            make_placeholder(
                outname.replace(".pdf", ""),
                outpath,
                note=f"Expected {csvn} or {xlsn} in {SRC}.",
            )
            print(f"Placeholder created for {outname}")
            continue
        rows = None
        if kind == "csv":
            rows = read_csv(path)
        else:
            rows = read_xlsx(path)
            if rows is None:
                make_placeholder(
                    outname.replace(".pdf", ""),
                    outpath,
                    note=f"XLSX found but openpyxl not installed: {path}",
                )
                print(f"Placeholder created for {outname} (openpyxl missing)")
                continue

        table, subtotal = build_table_rows(rows)
        try:
            make_pdf(outname.replace(".pdf", ""), table, subtotal, outpath)
            print(f"Exported: {outpath}")
        except Exception as e:
            print(f"Failed to create PDF for {outname}: {e}")


if __name__ == "__main__":
    main()
