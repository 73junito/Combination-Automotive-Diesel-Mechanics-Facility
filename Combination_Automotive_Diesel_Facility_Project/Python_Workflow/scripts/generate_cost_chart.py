"""
Generate TotalCost-per-bay bar chart and insert into portfolio Excel as a new sheet.
Saves chart image to outputs/totalcost_per_bay.png and appends sheet 'Cost Chart' to portfolio_equipment_by_bay.xlsx
"""

import os

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT = os.path.join(ROOT, "Python_Workflow", "outputs")
XLSX = os.path.join(OUT, "portfolio_equipment_by_bay.xlsx")
IMG_OUT = os.path.join(OUT, "totalcost_per_bay.png")


def main():
    if not os.path.exists(XLSX):
        raise FileNotFoundError(f"Excel file not found: {XLSX}")
    xls = pd.ExcelFile(XLSX)
    if "Bay Summary" not in xls.sheet_names:
        raise RuntimeError("Bay Summary sheet not found in portfolio Excel")
    summary = xls.parse("Bay Summary")
    # ensure numeric
    summary["TotalCost"] = pd.to_numeric(summary["TotalCost"], errors="coerce").fillna(
        0.0
    )
    # sort by TotalCost desc
    summary = summary.sort_values("TotalCost", ascending=False)
    # create bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(summary["BayName"].astype(str), summary["TotalCost"], color="tab:blue")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Total Cost")
    plt.title("Total Cost per Bay")
    plt.tight_layout()
    plt.savefig(IMG_OUT, dpi=150)
    plt.close()

    # insert into Excel as new sheet
    wb = load_workbook(XLSX)
    if "Cost Chart" in wb.sheetnames:
        # remove old sheet
        std = wb["Cost Chart"]
        wb.remove(std)
    ws = wb.create_sheet("Cost Chart")
    img = XLImage(IMG_OUT)
    # place image at cell A1
    ws.add_image(img, "A1")
    wb.save(XLSX)
    print("Wrote chart image:", IMG_OUT)
    print("Appended sheet Cost Chart to:", XLSX)


if __name__ == "__main__":
    main()
